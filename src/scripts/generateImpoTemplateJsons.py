from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook


PROJECT_ROOT = Path("/Users/fran/Documents/projects/pettransfer")
IMPO_ROOT = (
    PROJECT_ROOT
    / "api/src/database/drive_files/Modelos de cotizaciones IMPO VIGENTES (1)"
)
SOURCE_ROOT = (
    IMPO_ROOT / "3 Modelos de cotizaciones IMPO paises que sirve LATAM Pet Transport"
)
OUTPUT_ROOT = IMPO_ROOT / "json_templates"
MANIFEST_PATH = IMPO_ROOT / "impo_templates_manifest.json"

COUNTRY_MAP = {
    "ARGENTINA": "argentina",
    "BRASIL": "brasil",
    "CHILE": "chile",
    "COLOMBIA": "colombia",
    "COSTA RICA": "costa_rica",
    "ECUADOR": "ecuador",
    "HONDURAS": "honduras",
    "MIAMI": "miami",
    "MÉXICO": "mexico",
    "PANAMA": "panama",
    "PARAGUAY": "paraguay",
    "PERU": "peru",
    "URUGUAY": "uruguay",
}

COUNTRY_SHORT = {
    "argentina": "arg",
    "brasil": "bra",
    "chile": "chl",
    "colombia": "col",
    "costa_rica": "cri",
    "ecuador": "ecu",
    "honduras": "hnd",
    "miami": "mia",
    "mexico": "mex",
    "panama": "pan",
    "paraguay": "pry",
    "peru": "per",
    "uruguay": "ury",
}

LOCATION_CODES = {"BOG", "CLO", "GIG", "GRU", "VCP", "CUN", "MEX", "MIA", "SAP"}
QUOTE_SHEET_NAMES = {"cotizac", "hoja1", "cotizacion", "quotation"}
CALC_SHEET_NAMES = {"calculo", "cálculo", "calculo "}


@dataclass
class WorkbookData:
    sheets: dict[str, list[list[Any]]]
    source_type: str


def normalize_text(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def ascii_lower(text: str) -> str:
    return normalize_text(text).lower()


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace("\xa0", " ")
        value = re.sub(r"\s+", " ", value).strip()
        return value or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def row_values(row: list[Any]) -> list[Any]:
    cleaned = [clean_cell(value) for value in row]
    while cleaned and cleaned[0] is None:
        cleaned.pop(0)
    while cleaned and cleaned[-1] is None:
        cleaned.pop()
    return cleaned


def open_xlsx(path: Path) -> WorkbookData:
    workbook = load_workbook(path, data_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = [row_values(list(row)) for row in sheet.iter_rows(values_only=True)]
        sheets[sheet_name] = rows
    return WorkbookData(sheets=sheets, source_type="xlsx")


def xlrd_cell_value(book: xlrd.book.Book, sheet: xlrd.sheet.Sheet, row_idx: int, col_idx: int) -> Any:
    cell = sheet.cell(row_idx, col_idx)
    value = cell.value
    if cell.ctype == xlrd.XL_CELL_DATE:
        dt = xlrd.xldate_as_datetime(value, book.datemode)
        return dt.date().isoformat()
    return value


def open_xls(path: Path) -> WorkbookData:
    workbook = xlrd.open_workbook(path)
    sheets: dict[str, list[list[Any]]] = {}
    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)
        rows = []
        for row_idx in range(sheet.nrows):
            row = [
                xlrd_cell_value(workbook, sheet, row_idx, col_idx)
                for col_idx in range(sheet.ncols)
            ]
            rows.append(row_values(row))
        sheets[sheet_name] = rows
    return WorkbookData(sheets=sheets, source_type="xls")


def open_workbook_data(path: Path) -> WorkbookData | None:
    extension = path.suffix.lower()
    if extension == ".xlsx":
        return open_xlsx(path)
    if extension == ".xls":
        return open_xls(path)
    return None


def detect_variant(name: str) -> list[str]:
    lowered = ascii_lower(name)
    variants = []
    if "esp" in lowered:
        variants.append("esp")
    if "notice" in lowered:
        variants.append("notice")
    if "puppy" in lowered:
        variants.append("puppy")
    return variants or ["default"]


def detect_animal_count(name: str) -> int | None:
    lowered = ascii_lower(name)
    patterns = [
        r"(\d+)\s*(pet|pets|perros?)\b",
        r"\b(\d+)\b(?=\.[a-z0-9]+$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, lowered)
        if match:
            return int(match.group(1))
    return None


def detect_location(relative_path: Path, file_name: str) -> str | None:
    for part in relative_path.parts[:-1]:
        if part.upper() in LOCATION_CODES:
            return ascii_lower(part)
    lowered_name = ascii_lower(file_name)
    for code in LOCATION_CODES:
        if re.search(rf"\b{code.lower()}\b", lowered_name):
            return ascii_lower(code)
    return None


def situation_key(variants: list[str], animal_count: int | None) -> str:
    parts: list[str] = []
    if animal_count is not None:
        parts.append(f"{animal_count}_pet" if animal_count == 1 else f"{animal_count}_pets")
    for variant in variants:
        if variant != "default":
            parts.append(variant)
    return "_".join(parts) if parts else "default"


def json_file_name(country: str, location: str | None, animal_count: int | None, variants: list[str]) -> str:
    parts = ["cot", "impo", COUNTRY_SHORT[country]]
    if location:
        parts.append(location)
    if animal_count is not None:
        parts.append(str(animal_count))
    for variant in variants:
        if variant != "default":
            parts.append(variant)
    return "_".join(parts) + ".json"


def looks_numeric(value: Any) -> bool:
    return isinstance(value, (int, float))


def choose_quote_sheet(workbook: WorkbookData) -> tuple[str, list[list[Any]]]:
    for sheet_name, rows in workbook.sheets.items():
        if ascii_lower(sheet_name) in QUOTE_SHEET_NAMES:
            return sheet_name, rows
    for sheet_name, rows in workbook.sheets.items():
        if any(row for row in rows):
            return sheet_name, rows
    return next(iter(workbook.sheets.items()))


def choose_calc_sheet(workbook: WorkbookData) -> tuple[str, list[list[Any]]] | None:
    for sheet_name, rows in workbook.sheets.items():
        if ascii_lower(sheet_name) in CALC_SHEET_NAMES:
            return sheet_name, rows
    return None


def parse_metadata(rows: list[list[Any]]) -> tuple[dict[str, Any], int]:
    metadata: dict[str, Any] = {}
    label_map = {
        "customer": "customer",
        "customer:": "customer",
        "origin": "origin",
        "origin:": "origin",
        "from": "origin",
        "to": "destination",
        "destination": "destination",
        "quotation date": "quotation_date",
        "quotation date:": "quotation_date",
        "quoted": "quotation_date",
        "quoted ": "quotation_date",
        "arrival date": "arrival_date",
        "arrival date:": "arrival_date",
        "trip date": "trip_date",
        "# animals:": "animals",
        "# animals": "animals",
    }
    last_index = 0
    for index, row in enumerate(rows):
        if not row:
            continue
        first = row[0]
        if isinstance(first, str):
            normalized = ascii_lower(first).strip()
            if normalized == "usd":
                last_index = index
                break
            key = label_map.get(normalized)
            if key:
                metadata[key] = row[1] if len(row) > 1 else None
                last_index = index
                continue
            if re.match(r"^\d+$", str(first)):
                last_index = index - 1
                break
            if "import" in normalized and "pet" in normalized:
                metadata["title"] = first
                last_index = index
    return metadata, last_index


def parse_quoted_items(rows: list[list[Any]], start_index: int) -> tuple[list[dict[str, Any]], dict[str, Any] | None, int]:
    quoted_items: list[dict[str, Any]] = []
    total: dict[str, Any] | None = None
    index = start_index
    while index < len(rows):
        row = rows[index]
        if not row:
            index += 1
            continue
        first = row[0]
        label = None
        item_number = None
        amount = None
        note = None
        if len(row) >= 3 and looks_numeric(first) and isinstance(row[1], str):
            item_number = int(first)
            label = row[1]
            amount = row[2] if looks_numeric(row[2]) else row[2]
            note = row[3] if len(row) > 3 else None
        elif len(row) >= 2 and isinstance(first, str) and looks_numeric(row[1]):
            label = first
            amount = row[1]
            note = row[2] if len(row) > 2 else None
        elif isinstance(first, str) and ascii_lower(first).startswith("total"):
            amount = None
            note = None
            for value in row[1:]:
                if amount is None and looks_numeric(value):
                    amount = value
                elif value is not None:
                    note = value
            total = {
                "label": first,
                "amount": amount,
                "note": note,
            }
            index += 1
            break
        else:
            break
        quoted_items.append(
            {
                "item_number": item_number,
                "label": label,
                "amount": amount,
                "note": note,
            }
        )
        index += 1
    return quoted_items, total, index


def is_heading_row(row: list[Any]) -> bool:
    if not row:
        return False
    first = row[0]
    if not isinstance(first, str):
        return False
    lowered = ascii_lower(first)
    if lowered in {"comments", "conditions of contract", "contact:"}:
        return True
    if "important notice" in lowered:
        return True
    if re.match(r"^\d+\.", lowered):
        return True
    return False


def parse_body_sections(rows: list[list[Any]], start_index: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    descriptions: list[dict[str, Any]] = []
    notes: list[dict[str, Any]] = []
    current_description: dict[str, Any] | None = None
    current_note: dict[str, Any] | None = None

    for row in rows[start_index:]:
        if not row:
            continue
        first = row[0]
        if looks_numeric(first) and len(row) >= 2 and isinstance(row[1], str):
            current_description = {"item_number": int(first), "title": row[1], "paragraphs": []}
            descriptions.append(current_description)
            current_note = None
            continue
        if is_heading_row(row):
            current_note = {"title": row[0], "paragraphs": []}
            notes.append(current_note)
            current_description = None
            continue
        text = " ".join(str(value) for value in row if value is not None)
        if current_note is not None:
            current_note["paragraphs"].append(text)
        elif current_description is not None:
            current_description["paragraphs"].append(text)
        else:
            notes.append({"title": "general", "paragraphs": [text]})
    return descriptions, notes


def parse_calculation(rows: list[list[Any]]) -> dict[str, Any] | None:
    header_index = None
    currency_index = None
    for index, row in enumerate(rows):
        if row and isinstance(row[0], str) and ascii_lower(row[0]) == "costo":
            header_index = index
            currency_index = index + 1
            break
    if header_index is None or currency_index is None or currency_index >= len(rows):
        return None
    currencies = [value for value in rows[currency_index] if value is not None]
    items = []
    subtotals = []
    for row in rows[currency_index + 1 :]:
        if not row:
            continue
        label = row[0]
        if not isinstance(label, str):
            continue
        entry = {
            "label": label,
            "values": [value for value in row[1:] if value is not None],
        }
        if ascii_lower(label).startswith("sub total") or ascii_lower(label).startswith("total"):
            subtotals.append(entry)
        else:
            items.append(entry)
    return {"currencies": currencies, "items": items, "subtotals": subtotals}


def parse_template(path: Path, manifest_entry: dict[str, Any], country: str, location: str | None) -> dict[str, Any]:
    workbook = open_workbook_data(path)
    base = {
        "title": manifest_entry["file_name"].replace(".json", ""),
        "country": country,
        "location": location,
        "animal_count": manifest_entry.get("animal_count"),
        "variants": manifest_entry.get("variants", []),
        "source_template": {
            "file_name": path.name,
            "relative_path": manifest_entry["source_relative_path"],
            "extension": manifest_entry["source_extension"],
        },
    }
    if workbook is None:
        base["parser_status"] = "unsupported"
        base["quoted_items"] = []
        base["descriptions"] = []
        base["notes"] = []
        return base

    quote_sheet_name, quote_rows = choose_quote_sheet(workbook)
    metadata, metadata_index = parse_metadata(quote_rows)
    quoted_items, total, next_index = parse_quoted_items(quote_rows, metadata_index + 1)
    descriptions, notes = parse_body_sections(quote_rows, next_index)

    base["parser_status"] = "parsed"
    base["sheet_name"] = quote_sheet_name
    base["metadata"] = metadata
    base["quoted_items"] = quoted_items
    base["total"] = total
    base["descriptions"] = descriptions
    base["notes"] = notes

    calc_sheet = choose_calc_sheet(workbook)
    if calc_sheet is not None:
        calc_sheet_name, calc_rows = calc_sheet
        calculation = parse_calculation(calc_rows)
        if calculation:
            base["calculation"] = {
                "sheet_name": calc_sheet_name,
                **calculation,
            }
    return base


def build_manifest() -> dict[str, Any]:
    manifest: dict[str, Any] = {}
    for country_dir in sorted((path for path in SOURCE_ROOT.iterdir() if path.is_dir()), key=lambda p: p.name):
        country = COUNTRY_MAP[country_dir.name]
        country_entry: dict[str, Any] = {"title": country}
        templates: dict[str, Any] = {}
        locations: dict[str, Any] = {}
        files = sorted(
            path for path in country_dir.rglob("*") if path.is_file() and path.name != ".DS_Store"
        )

        for file_path in files:
            relative_source = file_path.relative_to(SOURCE_ROOT)
            relative_country = file_path.relative_to(country_dir)
            location = detect_location(relative_country, file_path.name)
            variants = detect_variant(file_path.name)
            animal_count = detect_animal_count(file_path.name)
            key = situation_key(variants, animal_count)
            output_name = json_file_name(country, location, animal_count, variants)

            entry = {
                "file_name": output_name,
                "relative_path": f"json_templates/{output_name}",
                "source_file_name": file_path.name,
                "source_relative_path": str(relative_source),
                "source_extension": file_path.suffix.lower().lstrip("."),
                "animal_count": animal_count,
                "variants": variants,
                "situation_key": key,
            }
            if location:
                entry["location"] = location
                locations.setdefault(location, {})[key] = entry
            else:
                templates[key] = entry

        if templates:
            country_entry["templates"] = templates
        if locations:
            country_entry["locations"] = locations
        manifest[country] = country_entry
    return manifest


def generate_template_jsons(manifest: dict[str, Any]) -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    for country, country_entry in manifest.items():
        for template in country_entry.get("templates", {}).values():
            source_path = SOURCE_ROOT / template["source_relative_path"]
            output_path = IMPO_ROOT / template["relative_path"]
            data = parse_template(source_path, template, country, None)
            output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        for location, templates in country_entry.get("locations", {}).items():
            for template in templates.values():
                source_path = SOURCE_ROOT / template["source_relative_path"]
                output_path = IMPO_ROOT / template["relative_path"]
                data = parse_template(source_path, template, country, location)
                output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    manifest = build_manifest()
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    generate_template_jsons(manifest)
    print(f"Manifest updated: {MANIFEST_PATH}")
    print(f"Template JSONs generated in: {OUTPUT_ROOT}")


if __name__ == "__main__":
    main()
